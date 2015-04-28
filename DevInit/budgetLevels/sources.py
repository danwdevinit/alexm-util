#!/usr/bin/env python

#Import system
import openpyxl
import csv
import re
from openpyxl import load_workbook
import sys, os
from optparse import OptionParser
import pdb

#Parse Options
parser = OptionParser()
#parser.add_option("-i", "--input", dest="input", default = "S:/Projects/Programme resources/Data/Data sets/Domestic Government Expenditure/Government budgets/Final data government finance_VA100415_originalnetlending.xlsx",
parser.add_option("-i", "--input", dest="input", default = "S:/Projects/Programme resources/Data/Data sets/Domestic Government Expenditure/Government budgets/Final data government finance_VA100415_zeronetlending.xlsx",
                help="Input file", metavar="FILE")
parser.add_option("-o", "--output", dest="output", default="./sources.csv",
                help="Output CSV file", metavar="FILE")
(options, args) = parser.parse_args()

#Unicode print
def uni(input):
    output = str(unicode(input).encode(sys.stdout.encoding, 'replace'))
    return output

#Import xlsx data
inPath = options.input
wb = load_workbook(filename = inPath)
sheets = wb.get_sheet_names()

sources = []
for sheet in sheets:
    ws = wb.get_sheet_by_name(name=sheet)
    country = uni(sheet)
    iso = ""
    yearCols = {}
    print('Reading sheet: '+country)
    for row in ws.iter_rows():
        colLen = len(row)
        if row[0].column=="A" and row[0].row==1:
            iso = uni(row[0].value)
        if uni(row[1].value).lower() == "year":
            for i in range(3,colLen):
                val = uni(row[i].value)
                col = row[i].column
                if str(val).lower()!='none':
                    yearCols[col]=val
        for cell in row:
            if cell.value:
                if uni(cell.value).strip().lower()=="source":
                    try:
                        comment = cell.comment
                    except:
                        comment = False
                    if comment:
                        year = yearCols[cell.column]
                        obj = {}
                        obj['Country']=country
                        obj['ISO3']=iso
                        obj['Year']=uni(year)
                        obj['Comment']=uni(comment.text)
                        sources.append(obj)
#Output results
print('Writing CSV...')
keys = sources[0].keys()
with open(options.output, 'wb') as output_file:
    dict_writer = csv.DictWriter(output_file, keys)
    dict_writer.writeheader()
    dict_writer.writerows(sources)
print('Done.')