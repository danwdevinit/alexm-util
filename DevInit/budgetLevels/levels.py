#!/usr/bin/env python

#Import system
import openpyxl
import csv
import re
import itertools
import operator
from operator import itemgetter
from difflib import SequenceMatcher
import json
from openpyxl import load_workbook
import sys, os
from optparse import OptionParser
import pdb

#Parse Options
parser = OptionParser()
parser.add_option("-i", "--input", dest="input",
                help="Input file", metavar="FILE")
parser.add_option("-o", "--output", dest="output", default="./results.csv",
                help="Output CSV file", metavar="FILE")
parser.add_option("-j", "--outputjson", dest="outputjson", default="./results.json",
                help="Output json file", metavar="FILE")
parser.add_option("-d", "--dict", dest="dict", default="./orgDict.json",
                help="orgDict JSON file", metavar="FILE")
(options, args) = parser.parse_args()

#Unicode print
def uni(input):
    try:
        output = float(unicode(input).encode(sys.stdout.encoding, 'replace'))
    except:
        output = re.sub(r'[^a-zA-Z0-9-_\s]', '',unicode(input).encode(sys.stdout.encoding, 'replace')).strip()
    return output

#Import xlsx data
inPath = options.input
try:
    wb = load_workbook(filename = inPath, use_iterators = True, data_only=True)
except:
    raise Exception("Input xlsx path required!")
sheets = wb.get_sheet_names()

#Define hierarchy
try:
    with open(options.dict, 'r') as f:
         orgDict = json.load(f)
except:
    orgDict = {}
flatData = []
hierData = {"name":"budget","children":[]}
for k in range(0,5):
    sheet = sheets[k]
#for sheet in sheets:
    ws = wb.get_sheet_by_name(name=sheet)
    rowIndex = 0
    oldNames = []
    names = []
    levels = []
    years = []
    types = []
    values = []
    country = uni(sheet)
    print('Reading sheet: '+country)
    for row in ws.iter_rows():
        names.append(uni(row[0].value))
        oldNames.append(uni(row[1].value))
        levels.append(uni(row[2].value))
        colLen = len(row)
        if uni(row[1].value).lower() == "year":
            for i in range(3,colLen):
                val = uni(row[i].value)
                if str(val).lower()!='none':
                    years.append(val)
        if uni(row[1].value).lower() == "type":
            for i in range(3,colLen):
                val = uni(row[i].value)
                types.append(val)
        if rowIndex>=5:
            rowValues = []
            for i in range(3,colLen):
                val = uni(row[i].value)
                rowValues.append(val)
            values.append(rowValues)
        rowIndex+=1
    currency = oldNames[1]
    iso = names[0]
    names = names[5:]
    levels = levels[5:]
    nameLen = len(names)
    yearLen = len(years)
    for i in range(0,nameLen):
        name = names[i]
        level = str(levels[i])
        levelSlug = level
        if level.lower().find('l0')>-1:
            for j in range(0,yearLen):
                item = {}
                year = years[j]
                yearType = types[j]
                item['iso'] = iso
                item['country'] = country
                item['currency'] = currency
                item['year'] = year
                item['type'] = yearType
                item['l1'] = name
                item['l2'] = ""
                item['l3'] = ""
                item['l4'] = ""
                item['l5'] = ""
                item['value'] = values[i][j] if str(values[i][j]).lower()!='none' else ""
                flatData.append(item)
        elif level!='none' and level!='None':
            for j in range(0,yearLen):
                item = {}
                year = years[j]
                yearType = types[j]
                try:
                    levelDict = orgDict[country][levelSlug]
                except:
                    print("Please define '"+level+"' in the sheet named '"+country+"':")
                    if country not in orgDict:
                        orgDict[country] = {}
                    orgDict[country][levelSlug] = {}
                    if levelSlug[0:2].lower()=="l1":
                        orgDict[country][levelSlug]['l1'] = str(raw_input('L1:')).strip()
                        orgDict[country][levelSlug]['l2'] = ""
                        orgDict[country][levelSlug]['l3'] = ""
                        orgDict[country][levelSlug]['l4'] = ""
                    elif levelSlug[0:2].lower()=="l2":
                        orgDict[country][levelSlug]['l1'] = str(raw_input('L1:')).strip()
                        orgDict[country][levelSlug]['l2'] = str(raw_input('L2:')).strip()
                        orgDict[country][levelSlug]['l3'] = ""
                        orgDict[country][levelSlug]['l4'] = ""
                    elif levelSlug[0:2].lower()=="l3":
                        orgDict[country][levelSlug]['l1'] = str(raw_input('L1:')).strip()
                        orgDict[country][levelSlug]['l2'] = str(raw_input('L2:')).strip()
                        orgDict[country][levelSlug]['l3'] = str(raw_input('L3:')).strip()
                        orgDict[country][levelSlug]['l4'] = ""
                    else:
                        orgDict[country][levelSlug]['l1'] = str(raw_input('L1:')).strip()
                        orgDict[country][levelSlug]['l2'] = str(raw_input('L2:')).strip()
                        orgDict[country][levelSlug]['l3'] = str(raw_input('L3:')).strip()
                        orgDict[country][levelSlug]['l4'] = str(raw_input('L4:')).strip()
                    levelDict = orgDict[country][levelSlug]
                    print('Writing orgDict...')
                    with open(options.dict, 'w') as output_file:
                        json.dump(orgDict,output_file,ensure_ascii=False,sort_keys=True,indent=2)
                    print('Done.')
                item['iso'] = iso
                item['country'] = country
                item['currency'] = currency
                item['year'] = year
                item['type'] = yearType
                item['l1'] = levelDict['l1']
                item['l2'] = name if level.lower().find('l1')>-1 else levelDict['l2']
                item['l3'] = name if level.lower().find('l2')>-1 else levelDict['l3']
                item['l4'] = name if level.lower().find('l3')>-1 else levelDict['l4']
                item['l5'] = name if level.lower().find('l4')>-1 else ""
                item['value'] = values[i][j] if str(values[i][j]).lower()!='none' else ""
                flatData.append(item)
    
#Build hierarchical data.
#def similar(a,b):
#    return SequenceMatcher(None,a,b).ratio()
parentModel = []
for item in flatData:
    if item['l5']!="":
        obj0 = {}
        obj0['name'] = item['l5']
        obj0['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']+"#"+item['l4']+"#"+item['l5']
        obj0['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']+"#"+item['l4']
        obj0['value'] = item['value'] if str(item['value']).lower()!='none' else ""
        parentModel.append(obj0)
        obj1 = {}
        obj1['name'] = item['l4']
        obj1['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']+"#"+item['l4']
        obj1['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']
        obj1['value'] = ""
        parentModel.append(obj1)
        obj2 = {}
        obj2['name'] = item['l3']
        obj2['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']
        obj2['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']
        obj2['value'] = ""
        parentModel.append(obj2)
        obj3 = {}
        obj3['name'] = item['l2']
        obj3['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']
        obj3['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj3['value'] = ""
        parentModel.append(obj3)
        obj4 = {}
        obj4['name'] = item['l1']
        obj4['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj4['parent'] = item['country']+"#"+str(int(item['year']))
        obj4['value'] = ""
        parentModel.append(obj4)
        obj5 = {}
        obj5['name'] = str(int(item['year']))
        obj5['id'] = item['country']+"#"+str(int(item['year']))
        obj5['parent'] = item['country']
        obj5['value'] = ""
        parentModel.append(obj5)
        obj6 = {}
        obj6['name'] = item['country']
        obj6['id'] = item['country']
        obj6['parent'] = ""
        obj6['value'] = ""
        parentModel.append(obj6)
    elif item['l4']!="":
        obj1 = {}
        obj1['name'] = item['l4']
        obj1['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']+"#"+item['l4']
        obj1['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']
        obj1['value'] = item['value'] if str(item['value']).lower()!='none' else ""
        parentModel.append(obj1)
        obj2 = {}
        obj2['name'] = item['l3']
        obj2['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']
        obj2['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']
        obj2['value'] = ""
        parentModel.append(obj2)
        obj3 = {}
        obj3['name'] = item['l2']
        obj3['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']
        obj3['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj3['value'] = ""
        parentModel.append(obj3)
        obj4 = {}
        obj4['name'] = item['l1']
        obj4['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj4['parent'] = item['country']+"#"+str(int(item['year']))
        obj4['value'] = ""
        parentModel.append(obj4)
        obj5 = {}
        obj5['name'] = str(int(item['year']))
        obj5['id'] = item['country']+"#"+str(int(item['year']))
        obj5['parent'] = item['country']
        obj5['value'] = ""
        parentModel.append(obj5)
        obj6 = {}
        obj6['name'] = item['country']
        obj6['id'] = item['country']
        obj6['parent'] = ""
        obj6['value'] = ""
        parentModel.append(obj6)
    elif item['l3']!="":
        obj2 = {}
        obj2['name'] = item['l3']
        obj2['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']
        obj2['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']
        obj2['value'] = item['value'] if str(item['value']).lower()!='none' else ""
        parentModel.append(obj2)
        obj3 = {}
        obj3['name'] = item['l2']
        obj3['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']
        obj3['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj3['value'] = ""
        parentModel.append(obj3)
        obj4 = {}
        obj4['name'] = item['l1']
        obj4['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj4['parent'] = item['country']+"#"+str(int(item['year']))
        obj4['value'] = ""
        parentModel.append(obj4)
        obj5 = {}
        obj5['name'] = str(int(item['year']))
        obj5['id'] = item['country']+"#"+str(int(item['year']))
        obj5['parent'] = item['country']
        obj5['value'] = ""
        parentModel.append(obj5)
        obj6 = {}
        obj6['name'] = item['country']
        obj6['id'] = item['country']
        obj6['parent'] = ""
        obj6['value'] = ""
        parentModel.append(obj6)
    elif item['l2']!="":
        obj3 = {}
        obj3['name'] = item['l2']
        obj3['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']
        obj3['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj3['value'] = item['value'] if str(item['value']).lower()!='none' else ""
        parentModel.append(obj3)
        obj4 = {}
        obj4['name'] = item['l1']
        obj4['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj4['parent'] = item['country']+"#"+str(int(item['year']))
        obj4['value'] = ""
        parentModel.append(obj4)
        obj5 = {}
        obj5['name'] = str(int(item['year']))
        obj5['id'] = item['country']+"#"+str(int(item['year']))
        obj5['parent'] = item['country']
        obj5['value'] = ""
        parentModel.append(obj5)
        obj6 = {}
        obj6['name'] = item['country']
        obj6['id'] = item['country']
        obj6['parent'] = ""
        obj6['value'] = ""
        parentModel.append(obj6)
    elif item['l1']!="":
        obj4 = {}
        obj4['name'] = item['l1']
        obj4['id'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj4['parent'] = item['country']+"#"+str(int(item['year']))
        obj4['value'] = item['value'] if str(item['value']).lower()!='none' else ""
        parentModel.append(obj4)
        obj5 = {}
        obj5['name'] = str(int(item['year']))
        obj5['id'] = item['country']+"#"+str(int(item['year']))
        obj5['parent'] = item['country']
        obj5['value'] = ""
        parentModel.append(obj5)
        obj6 = {}
        obj6['name'] = item['country']
        obj6['id'] = item['country']
        obj6['parent'] = ""
        obj6['value'] = ""
        parentModel.append(obj6)
#Remove exact duplicates
getvals = operator.itemgetter('id','parent')
parentModel.sort(key=getvals)
results = []
groups = []
for k, g in itertools.groupby(parentModel,getvals):
    groups.append(list(g))
for group in groups:
    sum = 0
    hasData = False
    obj = group[0]
    for item in group:
        try:
            sum+=float(item['value'])
            hasData = True
        except:
            sum+=0
    if hasData:
        obj['value'] = sum
    else:
        obj['value'] = ''
    results.append(obj)
parentModel[:]=results
#Remove near duplicates
#seen = set()
#result = []

syms = ['\\', '|', '/', '-']
bs = '\b'
spinIndex = 0
def spin():
    global spinIndex
    spinIndex = 0 if spinIndex>3 else spinIndex
    sym = syms[spinIndex]
    sys.stdout.write("\b%s" % sym)
    sys.stdout.flush()
    spinIndex+=1

def buildTree(parent,arr):
    spin()
    children = [i for i in parentModel if i['parent']==parent]
    arr[:] = children
    for child in arr:
        child['children'] = []
        buildTree(child['id'],child['children'])
        #Keep negatives?
        #if child['value']<0:
        #    child['value']=child['value']*-1
        if len(child['children'])==0:
            del child['children']
            if child['value']=="":
            #Give value to non-valued end nodes?
            #    child['value']=1
                child['value']=""
        else:
            if child['value']=="":
                del child['value']
    
sys.stdout.write("Building tree... This can take a while....")
buildTree("",hierData['children'])
sys.stdout.write('\nDone.\n')

#Output results
print('Writing CSV...')
#Enforce order
#keys = flatData[0].keys()
keys = ['iso','country','currency','year','type','l1','l2','l3','l4','l5','value']
with open(options.output, 'wb') as output_file:
    dict_writer = csv.DictWriter(output_file, keys)
    dict_writer.writeheader()
    dict_writer.writerows(flatData)
print('Done.')
print('Writing JSON...')
with open(options.outputjson, 'w') as output_file:
    json.dump(hierData,output_file,ensure_ascii=False,indent=2)
print('Done.')