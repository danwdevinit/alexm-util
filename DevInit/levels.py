#!/usr/bin/env python

#Import system
import openpyxl
import csv
import re
import itertools
import operator
from operator import itemgetter
from difflib import SequenceMatcher
import json
from openpyxl import load_workbook
import sys, os
from optparse import OptionParser
import pdb

#Parse Options
parser = OptionParser()
parser.add_option("-i", "--input", dest="input",
                help="Input file", metavar="FILE")
parser.add_option("-o", "--output", dest="output", default="./results.csv",
                help="Output CSV file", metavar="FILE")
parser.add_option("-j", "--outputjson", dest="outputjson", default="./results.json",
                help="Output json file", metavar="FILE")
(options, args) = parser.parse_args()

#Unicode print
def uni(input):
    try:
        output = float(unicode(input).encode(sys.stdout.encoding, 'replace'))
    except:
        output = re.sub(r'[^a-zA-Z0-9-\s]', '',unicode(input).encode(sys.stdout.encoding, 'replace')).lower()
    return output

#Import xlsx data
inPath = options.input
try:
    wb = load_workbook(filename = inPath, use_iterators = True)
except:
    raise Exception("Input xlsx path required!")
sheets = wb.get_sheet_names()

#Define hierarchy
orgDict = {}
orgDict['l2reven'] = {'l1':'total revenue and grants','l2':'revenue','l3':''}
orgDict['l3tax'] = {'l1':'total revenue and grants','l2':'revenue','l3':'tax'}
orgDict['l3grant'] = {'l1':'total revenue and grants','l2':'revenue','l3':'grants'}
orgDict['l3nonta'] = {'l1':'total revenue and grants','l2':'revenue','l3':'nontax'}
orgDict['l2expen'] = {'l1':'expenditures and net lending','l2':'expenditure','l3':''}
orgDict['l3recur'] = {'l1':'expenditures and net lending','l2':'expenditure','l3':'recurrent'}
orgDict['l3devel'] = {'l1':'expenditures and net lending','l2':'expenditure','l3':'development'}
orgDict['l3curre'] = {'l1':'expenditures and net lending','l2':'expenditure','l3':'current'}
orgDict['l3lendi'] = {'l1':'expenditures and net lending','l2':'expenditure','l3':'lending'}
orgDict['l3capit'] = {'l1':'expenditures and net lending','l2':'expenditure','l3':'capital'}
orgDict['l3priva'] = {'l1':'expenditures and net lending','l2':'expenditure','l3':'privatisation'}
orgDict['l3excep'] = {'l1':'expenditures and net lending','l2':'expenditure','l3':'exceptional'}
orgDict['l2finan'] = {'l1':'total financing','l2':'financing','l3':''}
orgDict['l3exter'] = {'l1':'total financing','l2':'financing','l3':'externalfinance'}
orgDict['l3domes'] = {'l1':'total financing','l2':'financing','l3':'domesticfinance'}
orgDict['l3forei'] = {'l1':'total financing','l2':'financing','l3':'foreignfinance'}
flatData = []
hierData = {}
#Fix it at first 5 for now
for i in range(0,6):
    sheet = sheets[i]
#for sheet in sheets:
    ws = wb.get_sheet_by_name(name=sheet)
    rowIndex = 0
    names = []
    levels = []
    years = []
    types = []
    values = []
    country = uni(sheet)
    #print('Reading sheet: '+country)
    for row in ws.iter_rows():
        names.append(uni(row[0].value))
        levels.append(uni(row[1].value))
        colLen = len(row)
        if uni(row[0].value) == "year":
            for i in range(2,colLen):
                val = uni(row[i].value)
                if val!='none':
                    years.append(val)
        if uni(row[0].value) == "type":
            for i in range(2,colLen):
                val = uni(row[i].value)
                if val!='none':
                    types.append(val)
        if rowIndex>=5:
            rowValues = []
            for i in range(2,colLen):
                val = uni(row[i].value)
                rowValues.append(val)
            values.append(rowValues)
        rowIndex+=1
    currency = names[1]
    names = names[5:]
    levels = levels[5:]
    nameLen = len(names)
    yearLen = len(years)
    for i in range(0,nameLen):
        name = names[i]
        level = levels[i]
        levelSlug = level[0:7]
        if level.find('l1')>-1:
            for j in range(0,yearLen):
                item = {}
                year = years[j]
                yearType = types[j]
                item['country'] = country
                item['currency'] = currency
                item['year'] = year
                item['type'] = yearType
                item['l1'] = name
                item['l2'] = ""
                item['l3'] = ""
                item['l4'] = ""
                item['value'] = values[i][j]
                flatData.append(item)
        elif level!='none':
            for j in range(0,yearLen):
                item = {}
                year = years[j]
                yearType = types[j]
                try:
                    levelDict = orgDict[levelSlug]
                except:
                    raise Exception("Please define '"+level+"' in the sheet named '"+country+"'")
                item['country'] = country
                item['currency'] = currency
                item['year'] = year
                item['type'] = yearType
                item['l1'] = levelDict['l1']
                item['l2'] = levelDict['l2']
                item['l3'] = name if level.find('l2')>-1 else levelDict['l3']
                item['l4'] = name if levelDict['l3']!="" else ""
                item['value'] = values[i][j]
                flatData.append(item)

#Build hierarchical data... turns out this is completely the wrong approach
#need to use IDs of some sort (concatenation of all ancestors)
def similar(a,b):
    return SequenceMatcher(None,a,b).ratio()
parentModel = []
for item in flatData:
    if item['l4']!="":
        obj1 = {}
        obj1['node'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']+"#"+item['l4']
        obj1['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']
        obj1['value'] = item['value'] if item['value']!='none' else ""
        parentModel.append(obj1)
        obj2 = {}
        obj2['node'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']
        obj2['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']
        obj2['value'] = ""
        parentModel.append(obj2)
        obj3 = {}
        obj3['node'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']
        obj3['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj3['value'] = ""
        parentModel.append(obj3)
        obj4 = {}
        obj4['node'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj4['parent'] = item['country']+"#"+str(int(item['year']))
        obj4['value'] = ""
        parentModel.append(obj4)
        obj5 = {}
        obj5['node'] = item['country']+"#"+str(int(item['year']))
        obj5['parent'] = item['country']
        obj5['value'] = ""
        parentModel.append(obj5)
        obj6 = {}
        obj6['node'] = item['country']
        obj6['parent'] = ""
        obj6['value'] = ""
        parentModel.append(obj6)
    elif item['l3']!="":
        obj2 = {}
        obj2['node'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']+"#"+item['l3']
        obj2['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']
        obj2['value'] = item['value'] if item['value']!='none' else ""
        parentModel.append(obj2)
        obj3 = {}
        obj3['node'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']
        obj3['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj3['value'] = ""
        parentModel.append(obj3)
        obj4 = {}
        obj4['node'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj4['parent'] = item['country']+"#"+str(int(item['year']))
        obj4['value'] = ""
        parentModel.append(obj4)
        obj5 = {}
        obj5['node'] = item['country']+"#"+str(int(item['year']))
        obj5['parent'] = item['country']
        obj5['value'] = ""
        parentModel.append(obj5)
        obj6 = {}
        obj6['node'] = item['country']
        obj6['parent'] = ""
        obj6['value'] = ""
        parentModel.append(obj6)
    elif item['l2']!="":
        obj3 = {}
        obj3['node'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']+"#"+item['l2']
        obj3['parent'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj3['value'] = item['value'] if item['value']!='none' else ""
        parentModel.append(obj3)
        obj4 = {}
        obj4['node'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj4['parent'] = item['country']+"#"+str(int(item['year']))
        obj4['value'] = ""
        parentModel.append(obj4)
        obj5 = {}
        obj5['node'] = item['country']+"#"+str(int(item['year']))
        obj5['parent'] = item['country']
        obj5['value'] = ""
        parentModel.append(obj5)
        obj6 = {}
        obj6['node'] = item['country']
        obj6['parent'] = ""
        obj6['value'] = ""
        parentModel.append(obj6)
    elif item['l1']!="":
        obj4 = {}
        obj4['node'] = item['country']+"#"+str(int(item['year']))+"#"+item['l1']
        obj4['parent'] = item['country']+"#"+str(int(item['year']))
        obj4['value'] = item['value'] if item['value']!='none' else ""
        parentModel.append(obj4)
        obj5 = {}
        obj5['node'] = item['country']+"#"+str(int(item['year']))
        obj5['parent'] = item['country']
        obj5['value'] = ""
        parentModel.append(obj5)
        obj6 = {}
        obj6['node'] = item['country']
        obj6['parent'] = ""
        obj6['value'] = ""
        parentModel.append(obj6)
#Remove exact duplicates
getvals = operator.itemgetter('node','parent')
parentModel.sort(key=getvals)
results = []
groups = []
for k, g in itertools.groupby(parentModel,getvals):
    groups.append(list(g))
for group in groups:
    sum = 0
    hasData = False
    obj = group[0]
    for item in group:
        try:
            sum+=float(item['value'])
            hasData = True
        except:
            sum+=0
    if hasData:
        obj['value'] = sum
    else:
        obj['value'] = ''
    results.append(obj)
parentModel[:]=results
#Remove near duplicates
seen = set()
result = []

def displayChildren(parent,level):
    children = [i for i in parentModel if i['parent']==parent]
    tab = "\t"
    for row in children:
        try:
            print(tab*level+str(row['node'].split("#")[-1:][0])+": "+str(row['value']))
        except:
            print(tab*level+str(row['node'].split("#")[-1:][0]))
        displayChildren(row['node'],level+1)
sys.stdout = open('results.txt','w')
displayChildren("",0)

#def findChildren(parent):
#    children = [i for i in parentModel if i['parent']==parent]
#    for row in children:
#        findChildren(row['node'])

#Output results
#print('Writing CSV...')
#Enforce order
#keys = flatData[0].keys()
keys = ['country','currency','year','type','l1','l2','l3','l4','value']
with open(options.output, 'wb') as output_file:
    dict_writer = csv.DictWriter(output_file, keys)
    dict_writer.writeheader()
    dict_writer.writerows(flatData)
#print('Done.')
#print('Writing JSON...')
with open(options.outputjson, 'w') as output_file:
    json.dump(parentModel,output_file,ensure_ascii=False)
#print('Done.')